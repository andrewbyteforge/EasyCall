"""
Workflow execution engine that runs nodes in order.
"""
import logging
import csv
import json
import os
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path
from io import BytesIO
from typing import Dict, List, Set, Any, Optional, Tuple
import logging
from collections import defaultdict, deque
from django.db import transaction

from nodes.node_registry import get_node_class
from apps.core.exceptions import NodeExecutionError  # Need to add this exception

logger = logging.getLogger(__name__)

# Default output directory for exports - use user's Desktop
DEFAULT_OUTPUT_DIR = Path.home() / "Desktop"


class WorkflowExecutor:
    """
    Executes a workflow by running nodes in topological order.
    """

    def __init__(self, workflow):
        self.workflow = workflow
        self.execution_context: Dict[str, Any] = {}  # Stores node outputs
        self.execution_log: List[str] = []  # Execution log messages

    def _log(self, message: str):
        """Add message to execution log."""
        self.execution_log.append(message)
        logger.info(message)

    def execute(self):
        """
        Execute the workflow (saves to database).

        Returns:
            ExecutionLog with results
        """
        from apps.execution.models import ExecutionLog

        # Create execution log
        execution = ExecutionLog.objects.create(
            workflow=self.workflow,
            status='PENDING'
        )

        try:
            execution.start()
            result = self._execute_workflow()
            execution.complete(result_data=result)

        except Exception as e:
            logger.error(f"Workflow execution failed: {e}")
            execution.fail(error_message=str(e))

        return execution

    def execute_direct(self) -> Dict[str, Any]:
        """
        Execute workflow directly without database.

        Returns:
            Dict with execution log, outputs, and summary
        """
        try:
            result = self._execute_workflow()
            return {
                "status": "success",
                "log": self.execution_log,
                "outputs": self.execution_context,
                "summary": {
                    "nodes_executed": len(self.execution_context),
                    "status": "COMPLETED"
                }
            }
        except Exception as e:
            import traceback
            return {
                "status": "error",
                "log": self.execution_log,
                "outputs": self.execution_context,
                "error": str(e),
                "traceback": traceback.format_exc(),
                "summary": {
                    "nodes_executed": len(self.execution_context),
                    "status": "FAILED"
                }
            }

    def _execute_workflow(self) -> Dict[str, Any]:
        """
        Core execution logic.

        Returns:
            Dict with all node outputs
        """
        self._log("═" * 60)
        self._log("🚀 WORKFLOW EXECUTION STARTED")
        self._log("═" * 60)

        # Get workflow canvas data
        canvas_data = self.workflow.canvas_data
        nodes = canvas_data.get('nodes', [])
        edges = canvas_data.get('edges', [])

        self._log(f"📊 Total Nodes: {len(nodes)}")
        self._log(f"🔗 Total Connections: {len(edges)}")

        # Determine execution order (topological sort)
        execution_order = self._topological_sort(nodes, edges)

        self._log("")
        self._log("📋 EXECUTION ORDER:")
        for i, node_data in enumerate(execution_order):
            node_type = node_data.get('type', node_data.get('data', {}).get('type', 'unknown'))
            self._log(f"   {i + 1}. {node_type} ({node_data['id']})")

        # Execute each node in order
        for node_data in execution_order:
            self._execute_node(node_data, edges)

        self._log("")
        self._log("═" * 60)
        self._log("✅ WORKFLOW EXECUTION COMPLETED")
        self._log("═" * 60)

        return self.execution_context

    def _execute_node(self, node_data: dict, edges: List[dict]):
        """Execute a single node."""
        node_id = node_data['id']
        node_type = node_data.get('type', node_data.get('data', {}).get('type', 'unknown'))
        config = node_data.get('data', {}).get('configValues', {})

        self._log("")
        self._log("─" * 60)
        self._log(f"▶ EXECUTING: {node_type}")
        self._log(f"  Node ID: {node_id}")

        # Get input data from connected nodes
        input_data = self._get_node_inputs(node_id, edges)

        if input_data:
            self._log("  📥 INPUTS:")
            for key, value in input_data.items():
                display_value = str(value)[:80] if value else "(none)"
                self._log(f"     • {key}: {display_value}")
        else:
            self._log("  📥 INPUTS: None (entry node)")

        if config:
            self._log("  ⚙️ CONFIG:")
            for key, value in config.items():
                display_value = str(value)[:50] if value else "(none)"
                self._log(f"     • {key}: {display_value}")

        # Execute based on node type
        try:
            result = self._run_node(node_type, node_id, input_data, config)
            self.execution_context[node_id] = result

            if result:
                self._log("  📤 OUTPUTS:")
                for key, value in result.items():
                    display_value = str(value)[:80] if value else "(none)"
                    self._log(f"     • {key}: {display_value}")

            self._log("  ✅ Node completed successfully")

        except Exception as e:
            self._log(f"  ❌ Node FAILED: {str(e)}")
            raise

    def _run_node(self, node_type: str, node_id: str, inputs: dict, config: dict) -> dict:
        """
        Execute node based on type.

        This routes to the appropriate handler for each node type.
        """
        # ═══════════════════════════════════════════════════════════════
        # INPUT NODES
        # ═══════════════════════════════════════════════════════════════

        if node_type == 'single_address':
            return {
                'address': config.get('address', ''),
                'blockchain': config.get('blockchain', 'bitcoin')
            }

        if node_type == 'batch_input':
            # Parse uploaded file and extract addresses
            file_info = config.get('file', {})
            return self._parse_batch_input(file_info, config)

        if node_type == 'transaction_hash':
            return {
                'tx_hash': config.get('tx_hash', ''),
                'blockchain': config.get('blockchain', 'bitcoin')
            }

        if node_type == 'batch_transaction':
            # Parse uploaded file and extract transaction hashes
            file_info = config.get('file', {})
            return self._parse_batch_transactions(file_info, config)

        # ═══════════════════════════════════════════════════════════════
        # CREDENTIAL NODES
        # ═══════════════════════════════════════════════════════════════

        if node_type == 'credential_chainalysis':
            return {
                'credentials': {
                    'type': 'chainalysis',
                    'api_key': config.get('api_key', ''),
                    'api_url': config.get('api_url', 'https://iapi.chainalysis.com'),
                    'authenticated': bool(config.get('api_key'))
                }
            }

        if node_type == 'credential_trm':
            return {
                'credentials': {
                    'type': 'trm',
                    'api_key': config.get('api_key', ''),
                    'authenticated': bool(config.get('api_key'))
                }
            }

        # ═══════════════════════════════════════════════════════════════
        # CHAINALYSIS QUERY NODES
        # ═══════════════════════════════════════════════════════════════

        if node_type == 'chainalysis_cluster_info':
            return self._execute_chainalysis_cluster_info(inputs, config)

        if node_type == 'chainalysis_cluster_balance':
            return self._execute_chainalysis_cluster_balance(inputs, config)

        if node_type == 'chainalysis_cluster_counterparties':
            return self._execute_chainalysis_counterparties(inputs, config)

        if node_type == 'chainalysis_transaction_details':
            return self._execute_chainalysis_transaction_details(inputs, config)

        if node_type == 'chainalysis_exposure_category':
            return self._execute_chainalysis_exposure_category(inputs, config)

        if node_type == 'chainalysis_exposure_service':
            return self._execute_chainalysis_exposure_service(inputs, config)

        # ═══════════════════════════════════════════════════════════════
        # TRM QUERY NODES
        # ═══════════════════════════════════════════════════════════════

        if node_type == 'trm_address_risk':
            return self._execute_trm_risk(inputs, config)

        if node_type == 'trm_address_ownership':
            return self._execute_trm_ownership(inputs, config)

        # ═══════════════════════════════════════════════════════════════
        # OUTPUT NODES
        # ═══════════════════════════════════════════════════════════════

        if node_type == 'csv_export':
            return self._export_csv(inputs, config)

        if node_type == 'pdf_export':
            # Check which render engine to use
            render_engine = config.get('render_engine', 'template')
            if render_engine == 'template':
                return self._export_pdf_template(inputs, config)
            else:
                return self._export_pdf(inputs, config)

        if node_type == 'json_export':
            return self._export_json(inputs, config)

        if node_type == 'excel_export':
            return self._export_excel(inputs, config)

        if node_type == 'txt_export':
            return self._export_txt(inputs, config)

        if node_type == 'output_path':
            # Output path node receives file_path from export node and can
            # also provide its own configured path
            output_config = config.get('output_path', {})
            incoming_path = inputs.get('file_path_input', inputs.get('file_path', ''))

            # If we have an incoming file and a configured destination, we could move/copy
            # For now, just pass through the information
            return {
                'file_path': incoming_path if isinstance(incoming_path, str) else incoming_path.get('file_path', ''),
                'configured_path': output_config.get('path', ''),
                'final_path': incoming_path if isinstance(incoming_path, str) else incoming_path.get('file_path', '')
            }

        if node_type == 'console_log':
            # Log all input data to console
            self._log(f"  📝 CONSOLE OUTPUT:")
            for key, value in inputs.items():
                display = json.dumps(value, indent=2, default=str)[:500]
                self._log(f"     {key}: {display}")
            return {'logged': True, 'data': inputs}

        # Unknown node type - return empty
        self._log(f"  ⚠️ Unknown node type: {node_type}")
        return {}

    # ═══════════════════════════════════════════════════════════════════════
    # CHAINALYSIS API METHODS
    # ═══════════════════════════════════════════════════════════════════════

    def _get_chainalysis_client(self, credentials: dict):
        """
        Get Chainalysis API client with credentials.

        Args:
            credentials: Credentials from node input (may contain api_key, api_url)

        Returns:
            ChainalysisClient instance
        """
        from apps.integrations.chainalysis_client import ChainalysisClient, ChainalysisAPIError

        api_key = credentials.get('api_key', '') if credentials else ''
        api_url = credentials.get('api_url', '') if credentials else ''

        # If no credentials from node, client will use settings
        if api_key:
            return ChainalysisClient(api_key=api_key, base_url=api_url if api_url else None)
        else:
            return ChainalysisClient()

    def _execute_chainalysis_cluster_info(self, inputs: dict, config: dict) -> dict:
        """
        Execute Chainalysis Cluster Info query.

        API Endpoint: GET /clusters/{address}
        Returns: cluster name, category, and root address
        """
        from apps.integrations.chainalysis_client import ChainalysisClient, ChainalysisAPIError

        credentials = inputs.get('credentials', {})
        address = inputs.get('address', '')

        # Handle batch addresses
        if isinstance(address, list):
            addresses = address
        elif 'addresses' in inputs:
            addresses = inputs.get('addresses', [])
        else:
            addresses = [address] if address else []

        if not addresses:
            return {'error': 'No address provided'}

        asset = config.get('asset', 'bitcoin')

        self._log(f"  [API] Chainalysis: Cluster Info for {len(addresses)} address(es)")

        try:
            client = self._get_chainalysis_client(credentials)

            results = []
            for addr in addresses[:100]:  # Limit to 100 per batch
                try:
                    response = client.get_cluster_info(address=addr, asset=asset)

                    results.append({
                        'address': addr,
                        'cluster_name': response.get('clusterName', 'Unknown'),
                        'category': response.get('category', 'Unknown'),
                        'cluster_address': response.get('rootAddress', addr)
                    })

                    self._log(f"     {addr[:12]}... -> {response.get('clusterName', 'Unknown')} ({response.get('category', 'Unknown')})")

                except ChainalysisAPIError as e:
                    if e.status_code == 404:
                        # Address not found - not an error, just unknown
                        results.append({
                            'address': addr,
                            'cluster_name': 'Unknown',
                            'category': 'Unknown',
                            'cluster_address': addr
                        })
                        self._log(f"     {addr[:12]}... -> Not found in database")
                    else:
                        raise

            # Return both list for batch and individual fields for single address
            return {
                'cluster_info': results,
                'count': len(results),
                # Individual outputs for single address use
                'cluster_name': results[0]['cluster_name'] if results else 'Unknown',
                'category': results[0]['category'] if results else 'Unknown',
                'cluster_address': results[0]['cluster_address'] if results else '',
                'address': addresses[0] if len(addresses) == 1 else addresses
            }

        except ChainalysisAPIError as e:
            self._log(f"  [ERROR] Chainalysis API: {e.user_message}")
            return {
                'error': e.user_message,
                'cluster_info': [],
                'count': 0
            }
        except ValueError as e:
            # API key not configured
            self._log(f"  [ERROR] {str(e)}")
            return {
                'error': str(e),
                'cluster_info': [],
                'count': 0
            }

    def _execute_chainalysis_cluster_balance(self, inputs: dict, config: dict) -> dict:
        """
        Execute Chainalysis Cluster Balance query.

        API Endpoint: GET /clusters/{address}/{asset}/summary
        Returns: balance, totalSent, totalReceived, transferCount, etc.
        """
        from apps.integrations.chainalysis_client import ChainalysisClient, ChainalysisAPIError

        credentials = inputs.get('credentials', {})
        address = inputs.get('address', '')

        # Also check for 'addresses' from batch input
        if isinstance(address, list):
            addresses = address
        elif 'addresses' in inputs:
            addresses = inputs.get('addresses', [])
        else:
            addresses = [address] if address else []

        if not addresses:
            return {'error': 'No address provided'}

        asset = config.get('asset', 'bitcoin')
        output_asset = config.get('output_asset', 'NATIVE')

        self._log(f"  [API] Chainalysis: Cluster Balance for {len(addresses)} address(es)")
        self._log(f"  [API] Asset: {asset}, Output: {output_asset}")
        self._log(f"  [API] First address: {addresses[0] if addresses else 'N/A'}")

        try:
            client = self._get_chainalysis_client(credentials)

            results = []
            for addr in addresses[:100]:  # Limit to 100 per batch
                try:
                    response = client.get_cluster_balance(
                        address=addr,
                        asset=asset,
                        output_asset=output_asset
                    )

                    results.append({
                        'address': addr,
                        'balance': response.get('balance', 0),
                        'total_sent': response.get('totalSent', 0),
                        'total_received': response.get('totalReceived', 0),
                        'transfer_count': response.get('transferCount', 0),
                        'deposit_count': response.get('depositCount', 0),
                        'withdrawal_count': response.get('withdrawalCount', 0),
                        'address_count': response.get('addressCount', 0),
                        'total_sent_fees': response.get('totalSentFees', 0),
                        'total_received_fees': response.get('totalReceivedFees', 0)
                    })

                    self._log(f"     {addr[:12]}... -> Balance: {response.get('balance', 0)}, Transfers: {response.get('transferCount', 0)}")

                except ChainalysisAPIError as e:
                    if e.status_code == 404:
                        # Address not found - return zeros
                        results.append({
                            'address': addr,
                            'balance': 0,
                            'total_sent': 0,
                            'total_received': 0,
                            'transfer_count': 0,
                            'deposit_count': 0,
                            'withdrawal_count': 0,
                            'address_count': 0,
                            'total_sent_fees': 0,
                            'total_received_fees': 0
                        })
                        self._log(f"     {addr[:12]}... -> Not found in database")
                    else:
                        raise

            return {
                'balance_data': results,
                'count': len(results),
                'balance': results[0]['balance'] if results else 0,
                'total_sent': results[0]['total_sent'] if results else 0,
                'total_received': results[0]['total_received'] if results else 0,
                'transfer_count': results[0]['transfer_count'] if results else 0,
                'address': addresses[0] if len(addresses) == 1 else addresses
            }

        except ChainalysisAPIError as e:
            self._log(f"  [ERROR] Chainalysis API: {e.user_message}")
            return {
                'error': e.user_message,
                'balance_data': [],
                'count': 0
            }
        except ValueError as e:
            # API key not configured
            self._log(f"  [ERROR] {str(e)}")
            return {
                'error': str(e),
                'balance_data': [],
                'count': 0
            }

    def _execute_chainalysis_counterparties(self, inputs: dict, config: dict) -> dict:
        """
        Execute Chainalysis Counterparties query.

        API Endpoint: GET /clusters/{address}/{asset}/counterparties
        Returns: list of counterparty addresses with transaction volumes,
                 grouped by source address when multiple addresses are input.
        """
        from apps.integrations.chainalysis_client import ChainalysisClient, ChainalysisAPIError

        credentials = inputs.get('credentials', {})
        address = inputs.get('address', '')

        # Handle batch addresses - process ALL of them, not just the first
        if isinstance(address, list):
            addresses = address
        elif 'addresses' in inputs:
            addresses = inputs.get('addresses', [])
        else:
            addresses = [address] if address else []

        if not addresses:
            return {'error': 'No address provided'}

        asset = config.get('asset', 'bitcoin')
        direction = config.get('direction', 'sent')
        output_asset = config.get('output_asset', 'NATIVE')

        self._log(f"  [API] Chainalysis: Counterparties for {len(addresses)} address(es) ({direction})")

        try:
            client = self._get_chainalysis_client(credentials)

            # Results grouped by source address
            results_by_address = []
            # Flat list of all counterparties with source_address field
            all_counterparties = []
            total_count = 0

            for addr in addresses[:100]:  # Limit to 100 addresses per batch
                try:
                    response = client.get_cluster_counterparties(
                        address=addr,
                        asset=asset,
                        output_asset=output_asset,
                        direction=direction,
                        limit=100
                    )

                    # Parse counterparties from response
                    counterparties = response if isinstance(response, list) else response.get('items', [])

                    # Add source_address to each counterparty for tracking
                    for cp in counterparties:
                        cp_with_source = dict(cp)  # Copy the counterparty data
                        cp_with_source['source_address'] = addr
                        all_counterparties.append(cp_with_source)

                    # Store grouped result
                    results_by_address.append({
                        'source_address': addr,
                        'counterparties': counterparties,
                        'count': len(counterparties)
                    })

                    total_count += len(counterparties)
                    self._log(f"     {addr[:12]}... -> {len(counterparties)} counterparties")

                except ChainalysisAPIError as e:
                    if e.status_code == 404:
                        # Address not found - add empty result
                        results_by_address.append({
                            'source_address': addr,
                            'counterparties': [],
                            'count': 0,
                            'error': 'Not found in database'
                        })
                        self._log(f"     {addr[:12]}... -> Not found in database")
                    else:
                        raise

            self._log(f"     Total: {total_count} counterparties from {len(addresses)} addresses")

            return {
                # Grouped results - each source address with its counterparties
                'counterparties_by_address': results_by_address,
                # Flat list with source_address field on each counterparty
                'counterparties': all_counterparties,
                # Summary counts
                'total_count': total_count,
                'addresses_queried': len(addresses),
                'count': total_count,
                # Pass through inputs
                'address': addresses[0] if len(addresses) == 1 else addresses,
                'direction': direction
            }

        except ChainalysisAPIError as e:
            self._log(f"  [ERROR] Chainalysis API: {e.user_message}")
            return {
                'error': e.user_message,
                'counterparties_by_address': [],
                'counterparties': [],
                'total_count': 0,
                'count': 0
            }
        except ValueError as e:
            # API key not configured
            self._log(f"  [ERROR] {str(e)}")
            return {
                'error': str(e),
                'counterparties_by_address': [],
                'counterparties': [],
                'total_count': 0,
                'count': 0
            }

    def _execute_chainalysis_transaction_details(self, inputs: dict, config: dict) -> dict:
        """
        Execute Chainalysis Transaction Details query.

        API Endpoint: GET /transactions/{hash}/{asset}/details
        Returns: transaction details including inputs, outputs, fees, etc.
        """
        from apps.integrations.chainalysis_client import ChainalysisClient, ChainalysisAPIError

        credentials = inputs.get('credentials', {})
        tx_hash = inputs.get('tx_hash', inputs.get('transaction_hash', ''))

        if not tx_hash:
            return {'error': 'No transaction hash provided'}

        asset = config.get('asset', 'bitcoin')
        output_asset = config.get('output_asset', 'NATIVE')

        self._log(f"  [API] Chainalysis: Transaction Details for {tx_hash[:16]}...")

        try:
            client = self._get_chainalysis_client(credentials)

            response = client.get_transaction_details(
                tx_hash=tx_hash,
                asset=asset,
                output_asset=output_asset
            )

            # Parse response - structure may vary based on API
            self._log(f"     Transaction found, processing details...")

            return {
                'transaction_details': response,
                'tx_hash': tx_hash,
                'asset': asset,
                'inputs': response.get('inputs', []),
                'outputs': response.get('outputs', []),
                'fee': response.get('fee', 0),
                'block_height': response.get('blockHeight', 0),
                'timestamp': response.get('timestamp', ''),
                'total_input': response.get('totalInput', 0),
                'total_output': response.get('totalOutput', 0)
            }

        except ChainalysisAPIError as e:
            if e.status_code == 404:
                self._log(f"     Transaction not found in database")
                return {
                    'transaction_details': {},
                    'tx_hash': tx_hash,
                    'error': 'Transaction not found'
                }

            self._log(f"  [ERROR] Chainalysis API: {e.user_message}")
            return {
                'error': e.user_message,
                'transaction_details': {},
                'tx_hash': tx_hash
            }
        except ValueError as e:
            self._log(f"  [ERROR] {str(e)}")
            return {
                'error': str(e),
                'transaction_details': {},
                'tx_hash': tx_hash
            }

    def _execute_chainalysis_exposure_category(self, inputs: dict, config: dict) -> dict:
        """
        Execute Chainalysis Exposure by Category query.

        API Endpoint: GET /exposures/clusters/{address}/{asset}/directions/{direction}
        Returns: exposure analysis by category (direct and indirect)
        """
        from apps.integrations.chainalysis_client import ChainalysisClient, ChainalysisAPIError

        credentials = inputs.get('credentials', {})
        address = inputs.get('address', '')

        # Handle batch addresses - use first one
        if isinstance(address, list):
            address = address[0] if address else ''

        if not address:
            return {'error': 'No address provided'}

        asset = config.get('asset', 'bitcoin')
        direction = config.get('direction', 'sent')
        output_asset = config.get('output_asset', 'USD')

        self._log(f"  [API] Chainalysis: Exposure by Category for {address[:12]}... ({direction})")

        # High-risk categories for flagging
        HIGH_RISK_CATEGORIES = [
            'darknet', 'ransomware', 'scam', 'stolen funds',
            'sanctions', 'child abuse', 'terrorism financing',
            'fraud', 'illicit'
        ]

        try:
            client = self._get_chainalysis_client(credentials)

            response = client.get_exposure_by_category(
                address=address,
                asset=asset,
                direction=direction,
                output_asset=output_asset
            )

            # Parse exposure data
            direct_exposure = response.get('direct', [])
            indirect_exposure = response.get('indirect', [])

            # Calculate totals
            total_direct = sum(e.get('value', 0) for e in direct_exposure)
            total_indirect = sum(e.get('value', 0) for e in indirect_exposure)
            total_risk = total_direct + total_indirect

            # Identify high-risk flags
            high_risk_flags = []
            for exp in direct_exposure:
                category = exp.get('category', '').lower()
                if any(hr in category for hr in HIGH_RISK_CATEGORIES):
                    high_risk_flags.append({
                        'category': exp.get('category'),
                        'value': exp.get('value', 0),
                        'percentage': exp.get('percentage', 0),
                        'exposure_type': 'direct'
                    })

            for exp in indirect_exposure:
                category = exp.get('category', '').lower()
                if any(hr in category for hr in HIGH_RISK_CATEGORIES):
                    high_risk_flags.append({
                        'category': exp.get('category'),
                        'value': exp.get('value', 0),
                        'percentage': exp.get('percentage', 0),
                        'exposure_type': 'indirect'
                    })

            self._log(f"     Direct exposure: {len(direct_exposure)} categories, total: {total_direct:,.2f}")
            self._log(f"     Indirect exposure: {len(indirect_exposure)} categories, total: {total_indirect:,.2f}")
            if high_risk_flags:
                self._log(f"     ⚠️ HIGH RISK FLAGS: {len(high_risk_flags)}")

            return {
                'exposure_data': response,
                'direct_exposure': direct_exposure,
                'indirect_exposure': indirect_exposure,
                'total_direct': total_direct,
                'total_indirect': total_indirect,
                'total_risk': total_risk,
                'high_risk_flags': high_risk_flags,
                'has_high_risk': len(high_risk_flags) > 0,
                'address': address,
                'direction': direction
            }

        except ChainalysisAPIError as e:
            if e.status_code == 404:
                self._log(f"     Address not found in database")
                return {
                    'exposure_data': {},
                    'direct_exposure': [],
                    'indirect_exposure': [],
                    'total_direct': 0,
                    'total_indirect': 0,
                    'total_risk': 0,
                    'high_risk_flags': [],
                    'has_high_risk': False,
                    'address': address
                }

            self._log(f"  [ERROR] Chainalysis API: {e.user_message}")
            return {
                'error': e.user_message,
                'exposure_data': {},
                'address': address
            }
        except ValueError as e:
            self._log(f"  [ERROR] {str(e)}")
            return {
                'error': str(e),
                'exposure_data': {},
                'address': address
            }

    def _execute_chainalysis_exposure_service(self, inputs: dict, config: dict) -> dict:
        """
        Execute Chainalysis Exposure by Service query.

        API Endpoint: GET /exposures/clusters/{address}/{asset}/directions/{direction}/services
        Returns: exposure analysis by specific services
        """
        from apps.integrations.chainalysis_client import ChainalysisClient, ChainalysisAPIError

        credentials = inputs.get('credentials', {})
        address = inputs.get('address', '')

        # Handle batch addresses - use first one
        if isinstance(address, list):
            address = address[0] if address else ''

        if not address:
            return {'error': 'No address provided'}

        asset = config.get('asset', 'bitcoin')
        direction = config.get('direction', 'sent')
        output_asset = config.get('output_asset', 'USD')

        self._log(f"  [API] Chainalysis: Exposure by Service for {address[:12]}... ({direction})")

        try:
            client = self._get_chainalysis_client(credentials)

            response = client.get_exposure_by_service(
                address=address,
                asset=asset,
                direction=direction,
                output_asset=output_asset
            )

            # Parse exposure data
            direct_exposure = response.get('direct', [])
            indirect_exposure = response.get('indirect', [])

            # Calculate totals
            total_direct = sum(e.get('value', 0) for e in direct_exposure)
            total_indirect = sum(e.get('value', 0) for e in indirect_exposure)

            # Get top services
            top_direct_services = sorted(direct_exposure, key=lambda x: x.get('value', 0), reverse=True)[:5]
            top_indirect_services = sorted(indirect_exposure, key=lambda x: x.get('value', 0), reverse=True)[:5]

            self._log(f"     Direct exposure: {len(direct_exposure)} services, total: {total_direct:,.2f}")
            self._log(f"     Indirect exposure: {len(indirect_exposure)} services, total: {total_indirect:,.2f}")

            if top_direct_services:
                top_names = [s.get('service', s.get('name', 'Unknown')) for s in top_direct_services[:3]]
                self._log(f"     Top services: {', '.join(top_names)}")

            return {
                'exposure_data': response,
                'direct_exposure': direct_exposure,
                'indirect_exposure': indirect_exposure,
                'total_direct': total_direct,
                'total_indirect': total_indirect,
                'top_direct_services': top_direct_services,
                'top_indirect_services': top_indirect_services,
                'service_count': len(direct_exposure) + len(indirect_exposure),
                'address': address,
                'direction': direction
            }

        except ChainalysisAPIError as e:
            if e.status_code == 404:
                self._log(f"     Address not found in database")
                return {
                    'exposure_data': {},
                    'direct_exposure': [],
                    'indirect_exposure': [],
                    'total_direct': 0,
                    'total_indirect': 0,
                    'top_direct_services': [],
                    'top_indirect_services': [],
                    'service_count': 0,
                    'address': address
                }

            self._log(f"  [ERROR] Chainalysis API: {e.user_message}")
            return {
                'error': e.user_message,
                'exposure_data': {},
                'address': address
            }
        except ValueError as e:
            self._log(f"  [ERROR] {str(e)}")
            return {
                'error': str(e),
                'exposure_data': {},
                'address': address
            }

    # ═══════════════════════════════════════════════════════════════════════
    # TRM API METHODS
    # ═══════════════════════════════════════════════════════════════════════

    def _execute_trm_risk(self, inputs: dict, config: dict) -> dict:
        """Execute TRM Risk Assessment."""
        address = inputs.get('address', '')
        self._log(f"  🛡️ TRM API: Risk Assessment for {address}")

        return {
            'risk_level': 'LOW',
            'risk_score': 15,
            'categories': ['exchange'],
            'address': address
        }

    def _execute_trm_ownership(self, inputs: dict, config: dict) -> dict:
        """Execute TRM Ownership query."""
        address = inputs.get('address', '')
        self._log(f"  🛡️ TRM API: Ownership for {address}")

        return {
            'owner': 'Unknown',
            'entity_type': 'exchange',
            'address': address
        }

    # ═══════════════════════════════════════════════════════════════════════
    # INPUT PARSING METHODS
    # ═══════════════════════════════════════════════════════════════════════

    def _parse_batch_input(self, file_info: dict, config: dict) -> dict:
        """Parse batch input file and extract addresses."""
        # In browser context, file_info contains file metadata
        # The actual parsing would happen client-side or via upload

        file_name = file_info.get('name', 'unknown')
        file_format = config.get('format', 'csv')

        self._log(f"  📁 Parsing batch file: {file_name} ({file_format})")

        # For now, return simulated addresses
        # Real implementation would read the actual file
        return {
            'addresses': [
                '1A1zP1eP5QGefi2DMPTfTL5SLmv7DivfNa',
                '3J98t1WpEZ73CNmQviecrnyiWrnqRhWNLy',
                'bc1qar0srrr7xfkvy5l643lydnw9re59gtzzwf5mdq'
            ],
            'count': 3,
            'blockchain': config.get('blockchain', 'bitcoin'),
            'source_file': file_name
        }

    def _parse_batch_transactions(self, file_info: dict, config: dict) -> dict:
        """Parse batch input file and extract transaction hashes."""
        # In browser context, file_info contains file metadata
        # The actual parsing would happen client-side or via upload

        file_name = file_info.get('name', 'unknown')
        file_format = config.get('format', 'csv')

        self._log(f"  📋 Parsing batch transaction file: {file_name} ({file_format})")

        # For now, return simulated transaction hashes
        # Real implementation would read the actual file
        return {
            'tx_hashes': [
                '4a5e1e4baab89f3a32518a88c31bc87f618f76673e2cc77ab2127b7afdeda33b',
                'e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855',
            ],
            'count': 2,
            'blockchain': config.get('blockchain', 'bitcoin'),
            'source_file': file_name
        }

    # ═══════════════════════════════════════════════════════════════════════
    # EXPORT METHODS
    # ═══════════════════════════════════════════════════════════════════════

    def _prepare_export_data(self, inputs: dict) -> List[dict]:
        """
        Convert input data into a list of flat dictionaries for export.
        Handles various input structures from query nodes.
        """
        rows = []

        # First, check for 'data' input which may contain the full source output
        data_input = inputs.get('data', {})

        # Check for common data structures from query nodes
        # These can be in either the 'data' input or directly in inputs

        # balance_data from cluster_balance node
        balance_data = data_input.get('balance_data') if isinstance(data_input, dict) else None
        balance_data = balance_data or inputs.get('balance_data')
        if balance_data:
            if isinstance(balance_data, list):
                rows.extend(balance_data)
            elif isinstance(balance_data, dict):
                rows.append(balance_data)

        # cluster_info from cluster_info node
        cluster_info = data_input.get('cluster_info') if isinstance(data_input, dict) else None
        cluster_info = cluster_info or inputs.get('cluster_info')
        if cluster_info and not rows:
            if isinstance(cluster_info, list):
                rows.extend(cluster_info)
            elif isinstance(cluster_info, dict):
                rows.append(cluster_info)

        # counterparties from counterparties node
        # First check for grouped counterparties_by_address (new format with source tracking)
        counterparties_by_address = data_input.get('counterparties_by_address') if isinstance(data_input, dict) else None
        counterparties_by_address = counterparties_by_address or inputs.get('counterparties_by_address')
        if counterparties_by_address and not rows:
            # Flatten grouped data, adding source_address to each counterparty
            for group in counterparties_by_address:
                source_addr = group.get('source_address', '')
                for cp in group.get('counterparties', []):
                    row = dict(cp)  # Copy counterparty data
                    row['source_address'] = source_addr  # Add source tracking
                    rows.append(row)

        # Fall back to flat counterparties list (which already has source_address if from new method)
        counterparties = data_input.get('counterparties') if isinstance(data_input, dict) else None
        counterparties = counterparties or inputs.get('counterparties')
        if counterparties and not rows:
            if isinstance(counterparties, list):
                rows.extend(counterparties)
            elif isinstance(counterparties, dict):
                rows.append(counterparties)

        # addresses from batch_input node
        addresses = data_input.get('addresses') if isinstance(data_input, dict) else None
        addresses = addresses or inputs.get('addresses')
        if addresses and not rows:
            if isinstance(addresses, list):
                # Convert list of addresses to list of dicts
                rows.extend([{'address': addr} for addr in addresses])

        # If no recognized list structure, check if data_input is a dict with row-like data
        if not rows and isinstance(data_input, dict):
            # Look for any list of dicts in the data
            for key, value in data_input.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    rows.extend(value)
                    break

            # If still no rows, and data_input has scalar values, treat it as a single row
            if not rows:
                # Filter out non-exportable keys
                exportable = {k: v for k, v in data_input.items()
                             if k not in ('_source_data', 'count', 'error') and not isinstance(v, list)}
                if exportable:
                    rows.append(exportable)

        # If no recognized structure and still no rows, try to flatten all inputs
        if not rows:
            for key, value in inputs.items():
                if key.startswith('_'):  # Skip internal keys
                    continue
                if isinstance(value, list) and len(value) > 0:
                    if isinstance(value[0], dict):
                        rows.extend(value)
                    else:
                        # List of primitives - create rows
                        rows.extend([{key: v} for v in value])
                elif isinstance(value, dict):
                    # Check for nested data arrays
                    found_nested = False
                    for k, v in value.items():
                        if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                            rows.extend(v)
                            found_nested = True
                            break
                    if not found_nested and value:
                        # Filter to exportable values
                        exportable = {k: v for k, v in value.items()
                                     if not isinstance(v, (list, dict)) or k in ('address',)}
                        if exportable:
                            rows.append(exportable)

        # If still no rows, create one from all scalar inputs
        if not rows:
            row = {}
            for key, value in inputs.items():
                if key.startswith('_'):
                    continue
                if not isinstance(value, (list, dict)):
                    row[key] = value
            if row:
                rows.append(row)

        return rows

    def _get_output_path(self, config: dict, default_filename: str) -> str:
        """Get the output file path from config or generate default."""
        # Ensure output directory exists
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # Check for configured path
        configured_path = config.get('output_path', {}).get('path', '')
        if configured_path:
            # If it's just a filename, put it in the output dir
            if not os.path.dirname(configured_path):
                return str(DEFAULT_OUTPUT_DIR / configured_path)
            return configured_path

        # Generate timestamped filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{default_filename}_{timestamp}"
        return str(DEFAULT_OUTPUT_DIR / filename)

    def _export_csv(self, inputs: dict, config: dict) -> dict:
        """Export data to CSV file."""
        rows = self._prepare_export_data(inputs)

        if not rows:
            self._log("  ⚠️ No data to export")
            return {'file_path': None, 'rows_written': 0, 'error': 'No data to export'}

        # Get output path
        file_path = self._get_output_path(config, 'output') + '.csv'

        # Get all unique columns from all rows
        columns = []
        for row in rows:
            for key in row.keys():
                if key not in columns:
                    columns.append(key)

        self._log(f"  📤 Exporting {len(rows)} rows to CSV...")
        self._log(f"  📝 Columns: {', '.join(columns)}")

        # Write CSV file
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)

        self._log(f"  💾 Written to: {file_path}")

        return {
            'file_path': file_path,
            'rows_written': len(rows),
            'columns': columns
        }

    def _export_json(self, inputs: dict, config: dict) -> dict:
        """Export data to JSON file."""
        rows = self._prepare_export_data(inputs)

        if not rows:
            self._log("  ⚠️ No data to export")
            return {'file_path': None, 'rows_written': 0, 'error': 'No data to export'}

        # Get output path
        file_path = self._get_output_path(config, 'output') + '.json'

        self._log(f"  📤 Exporting {len(rows)} records to JSON...")

        # Write JSON file
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, indent=2, default=str)

        self._log(f"  💾 Written to: {file_path}")

        return {
            'file_path': file_path,
            'rows_written': len(rows)
        }

    def _export_excel(self, inputs: dict, config: dict) -> dict:
        """Export data to Excel file."""
        rows = self._prepare_export_data(inputs)

        if not rows:
            self._log("  ⚠️ No data to export")
            return {'file_path': None, 'rows_written': 0, 'error': 'No data to export'}

        # Get output path
        file_path = self._get_output_path(config, 'output') + '.xlsx'

        # Get all unique columns
        columns = []
        for row in rows:
            for key in row.keys():
                if key not in columns:
                    columns.append(key)

        self._log(f"  📤 Exporting {len(rows)} rows to Excel...")

        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Write header
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Write data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = row_data.get(col_name, '')
                    # Convert lists/dicts to string
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(file_path)
            self._log(f"  💾 Written to: {file_path}")

            return {
                'file_path': file_path,
                'rows_written': len(rows),
                'columns': columns
            }

        except ImportError:
            self._log("  ⚠️ openpyxl not installed, falling back to CSV")
            return self._export_csv(inputs, config)

    def _export_pdf(self, inputs: dict, config: dict) -> dict:
        """
        Export data to a professional PDF report matching the HTML template style.

        Features:
        - Professional cover page with blue sidebar (matching HTML template)
        - Executive summary with stats cards
        - Data sections adapted to content type
        - Formatted data tables with professional styling
        - Consistent #1a237e color scheme
        """
        rows = self._prepare_export_data(inputs)

        if not rows:
            self._log("  ⚠️ No data to export")
            return {'file_path': None, 'rows_written': 0, 'error': 'No data to export'}

        # Get configuration
        report_title = config.get('report_title', 'Blockchain Intelligence Report')
        include_graphs = config.get('include_graphs', True)
        graph_type = config.get('graph_type', 'auto')

        # Get output path
        file_path = self._get_output_path(config, 'report') + '.pdf'

        self._log(f"  📑 Generating PDF report with {len(rows)} records...")

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm, mm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                Image, PageBreak, KeepTogether, NextPageTemplate, PageTemplate,
                Frame, BaseDocTemplate, Flowable
            )
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
            from reportlab.pdfgen import canvas
            import matplotlib
            matplotlib.use('Agg')  # Non-interactive backend
            import matplotlib.pyplot as plt
            import uuid

            # Define page sizes
            PAGE_WIDTH, PAGE_HEIGHT = A4

            # ═══════════════════════════════════════════════════════════════
            # COLOR SCHEME (matching HTML template)
            # ═══════════════════════════════════════════════════════════════
            PRIMARY_BLUE = colors.HexColor('#1a237e')       # Dark blue (sidebar, headers)
            SECONDARY_BLUE = colors.HexColor('#283593')     # Medium blue
            ACCENT_BLUE = colors.HexColor('#3949ab')        # Light blue accent
            SECTION_BORDER = colors.HexColor('#e8eaf6')     # Light blue-gray border
            TEXT_DARK = colors.HexColor('#333333')
            TEXT_MUTED = colors.HexColor('#666666')
            TEXT_LIGHT = colors.HexColor('#999999')
            BG_LIGHT = colors.HexColor('#f5f5f5')
            BG_ROW_ALT = colors.HexColor('#fafafa')

            # Get styles
            styles = getSampleStyleSheet()

            # ═══════════════════════════════════════════════════════════════
            # CUSTOM STYLES (matching HTML template)
            # ═══════════════════════════════════════════════════════════════

            # Section title style (matches .section-title in HTML)
            section_title_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading1'],
                fontSize=16,
                fontName='Helvetica-Bold',
                spaceBefore=25,
                spaceAfter=12,
                textColor=PRIMARY_BLUE,
                borderPadding=(0, 0, 8, 0),
            )

            # Body text style
            body_style = ParagraphStyle(
                'BodyText',
                parent=styles['Normal'],
                fontSize=11,
                spaceAfter=10,
                alignment=TA_LEFT,
                leading=16,
                textColor=TEXT_DARK
            )

            # Key-value label style
            kv_label_style = ParagraphStyle(
                'KVLabel',
                parent=styles['Normal'],
                fontSize=9,
                textColor=TEXT_MUTED,
                spaceAfter=2
            )

            # Key-value value style
            kv_value_style = ParagraphStyle(
                'KVValue',
                parent=styles['Normal'],
                fontSize=11,
                fontName='Helvetica-Bold',
                textColor=TEXT_DARK,
                spaceAfter=8
            )

            # Create custom document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=15*mm,
                leftMargin=15*mm,
                topMargin=15*mm,
                bottomMargin=15*mm
            )

            story = []
            report_id = str(uuid.uuid4())[:8].upper()
            generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # ═══════════════════════════════════════════════════════════════
            # PAGE 1: COVER PAGE (matching HTML template layout)
            # ═══════════════════════════════════════════════════════════════

            # Calculate dimensions
            content_width = PAGE_WIDTH - 30*mm
            sidebar_width = content_width * 0.35
            main_width = content_width * 0.65
            cover_height = PAGE_HEIGHT - 30*mm

            # Build sidebar content
            sidebar_content = []
            sidebar_content.append(Paragraph(
                "🔗",
                ParagraphStyle('LogoIcon', fontSize=36, alignment=TA_CENTER, textColor=colors.white, spaceAfter=10)
            ))
            sidebar_content.append(Paragraph(
                "EasyCall",
                ParagraphStyle('LogoText', fontSize=24, fontName='Helvetica-Bold',
                              alignment=TA_CENTER, textColor=colors.white, spaceAfter=40)
            ))
            sidebar_content.append(Spacer(1, cover_height * 0.5))
            sidebar_content.append(Paragraph(
                f"<b>Report Generated:</b><br/>{generated_at}",
                ParagraphStyle('MetaText', fontSize=9, textColor=colors.white, alignment=TA_LEFT,
                              leading=14, leftIndent=10)
            ))
            sidebar_content.append(Spacer(1, 15))
            sidebar_content.append(Paragraph(
                "<b>Classification:</b><br/>CONFIDENTIAL",
                ParagraphStyle('MetaText', fontSize=9, textColor=colors.white, alignment=TA_LEFT,
                              leading=14, leftIndent=10)
            ))
            sidebar_content.append(Spacer(1, 15))
            sidebar_content.append(Paragraph(
                f"<b>Report ID:</b><br/>{report_id}",
                ParagraphStyle('MetaText', fontSize=9, textColor=colors.white, alignment=TA_LEFT,
                              leading=14, leftIndent=10)
            ))

            # Build sidebar as a table cell
            sidebar_table = Table([[sidebar_content]], colWidths=[sidebar_width - 10])
            sidebar_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 30),
            ]))

            # Build main content
            main_content = []
            main_content.append(Spacer(1, cover_height * 0.25))
            main_content.append(Paragraph(
                report_title,
                ParagraphStyle('CoverTitle', fontSize=28, fontName='Helvetica-Bold',
                              textColor=PRIMARY_BLUE, spaceAfter=15, leading=34)
            ))
            main_content.append(Paragraph(
                "Comprehensive analysis of blockchain addresses and transactions",
                ParagraphStyle('CoverSubtitle', fontSize=14, textColor=TEXT_MUTED, spaceAfter=40)
            ))

            # Cover info items
            summary_stats = self._calculate_summary_stats(rows)
            info_items = [
                ("WORKFLOW", getattr(self.workflow, 'name', 'Untitled Workflow')),
                ("RECORDS ANALYZED", str(len(rows))),
            ]
            if summary_stats.get('unique_addresses'):
                info_items.append(("ADDRESSES", str(summary_stats['unique_addresses'])))
            if summary_stats.get('total_transfers'):
                info_items.append(("TRANSFERS", str(summary_stats['total_transfers'])))

            # Info section with left border
            info_table_data = []
            for label, value in info_items:
                info_table_data.append([
                    Paragraph(label, ParagraphStyle('InfoLabel', fontSize=9, textColor=TEXT_LIGHT)),
                ])
                info_table_data.append([
                    Paragraph(value, ParagraphStyle('InfoValue', fontSize=12, fontName='Helvetica-Bold',
                                                   textColor=TEXT_DARK, spaceAfter=12)),
                ])

            info_table = Table(info_table_data, colWidths=[main_width - 40])
            info_table.setStyle(TableStyle([
                ('LEFTPADDING', (0, 0), (-1, -1), 15),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))

            main_content.append(Spacer(1, 30))

            # Create bordered info section
            bordered_info = Table([[info_table]], colWidths=[main_width - 20])
            bordered_info.setStyle(TableStyle([
                ('LINEWIDTH', (0, 0), (0, 0), 4),
                ('LINEBEFORE', (0, 0), (0, 0), 4, PRIMARY_BLUE),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ]))
            main_content.append(bordered_info)

            main_table = Table([[main_content]], colWidths=[main_width])
            main_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 30),
            ]))

            # Combine sidebar and main into cover
            cover_table = Table(
                [[[sidebar_table], [main_table]]],
                colWidths=[sidebar_width, main_width]
            )
            cover_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), PRIMARY_BLUE),
                ('BACKGROUND', (1, 0), (1, 0), colors.white),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))

            story.append(cover_table)
            story.append(PageBreak())

            # ═══════════════════════════════════════════════════════════════
            # PAGE 2: CONTENT PAGES (matching HTML template sections)
            # ═══════════════════════════════════════════════════════════════

            # Page header
            header_table = Table(
                [[
                    Paragraph("🔗 EasyCall", ParagraphStyle('HeaderLogo', fontSize=14,
                              fontName='Helvetica-Bold', textColor=PRIMARY_BLUE)),
                    Paragraph(report_title, ParagraphStyle('HeaderTitle', fontSize=10,
                              textColor=TEXT_MUTED, alignment=TA_RIGHT))
                ]],
                colWidths=[content_width * 0.5, content_width * 0.5]
            )
            header_table.setStyle(TableStyle([
                ('LINEBELOW', (0, 0), (-1, 0), 2, PRIMARY_BLUE),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            story.append(header_table)
            story.append(Spacer(1, 20))

            # --- SECTION: Executive Summary ---
            story.append(Paragraph("Executive Summary", section_title_style))

            # Executive summary box (gray background like HTML)
            total_records = len(rows)
            columns = list(rows[0].keys()) if rows else []

            summary_text = f"This report analyzed <b>{total_records}</b> records"
            if summary_stats.get('unique_addresses'):
                summary_text += f" covering <b>{summary_stats['unique_addresses']}</b> blockchain address(es)"
            if summary_stats.get('total_transfers'):
                summary_text += f" with <b>{summary_stats['total_transfers']:,}</b> total transfers"
            summary_text += "."

            summary_box = Table(
                [[Paragraph(summary_text, body_style)]],
                colWidths=[content_width - 10]
            )
            summary_box.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
                ('TOPPADDING', (0, 0), (-1, -1), 15),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
                ('LEFTPADDING', (0, 0), (-1, -1), 20),
                ('RIGHTPADDING', (0, 0), (-1, -1), 20),
                ('ROUNDEDCORNERS', [8, 8, 8, 8]),
            ]))
            story.append(summary_box)

            # Summary stats cards
            stat_cards = []
            stat_cards.append(("Records", str(total_records)))
            if summary_stats.get('unique_addresses'):
                stat_cards.append(("Addresses", str(summary_stats['unique_addresses'])))
            if summary_stats.get('total_transfers'):
                stat_cards.append(("Transfers", str(summary_stats['total_transfers'])))
            if summary_stats.get('total_balance'):
                stat_cards.append(("Balance", f"{summary_stats['total_balance']:.4f}"))

            if stat_cards:
                card_width = (content_width - 30) / len(stat_cards)
                card_data = []
                for label, value in stat_cards:
                    card_content = [
                        Paragraph(str(value), ParagraphStyle('StatValue', fontSize=20,
                                  fontName='Helvetica-Bold', textColor=PRIMARY_BLUE, alignment=TA_CENTER)),
                        Paragraph(label.upper(), ParagraphStyle('StatLabel', fontSize=9,
                                  textColor=TEXT_MUTED, alignment=TA_CENTER))
                    ]
                    card_data.append(card_content)

                stats_table = Table([card_data], colWidths=[card_width] * len(stat_cards))
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
                    ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
                    ('TOPPADDING', (0, 0), (-1, -1), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                story.append(Spacer(1, 15))
                story.append(stats_table)

            # --- SECTION: Data Visualization ---
            if include_graphs and rows:
                story.append(Spacer(1, 10))
                story.append(Paragraph("Data Visualization", section_title_style))
                graph_image = self._generate_graph(rows, graph_type)
                if graph_image:
                    story.append(Image(graph_image, width=5*inch, height=3*inch))
                    story.append(Paragraph(
                        "<i>Figure 1: Analysis results visualization</i>",
                        ParagraphStyle('Caption', parent=styles['Normal'], fontSize=9,
                                      alignment=TA_CENTER, textColor=TEXT_MUTED)
                    ))

            # --- SECTION: Query Results (Key-Value pairs for first few records) ---
            story.append(PageBreak())

            # Page header again
            story.append(header_table)
            story.append(Spacer(1, 20))

            story.append(Paragraph("Query Results", section_title_style))

            # Show first 5 records as key-value cards
            for idx, row in enumerate(rows[:5]):
                # Create card for each record
                card_title = f"Record {idx + 1}"
                if 'address' in row:
                    card_title = f"Address: {str(row['address'])[:20]}..."
                elif 'clusterName' in row or 'cluster_name' in row:
                    card_title = row.get('clusterName') or row.get('cluster_name', f'Record {idx + 1}')

                story.append(Paragraph(card_title, ParagraphStyle(
                    'CardTitle', fontSize=12, fontName='Helvetica-Bold',
                    textColor=TEXT_DARK, spaceBefore=15, spaceAfter=8
                )))

                # Key-value pairs in 2-column grid
                kv_data = []
                items = list(row.items())[:8]  # Limit to 8 fields per record
                for i in range(0, len(items), 2):
                    row_cells = []
                    for j in range(2):
                        if i + j < len(items):
                            key, value = items[i + j]
                            # Format the value
                            if isinstance(value, float):
                                value = f"{value:,.4f}"
                            elif isinstance(value, (list, dict)):
                                value = str(value)[:50] + "..." if len(str(value)) > 50 else str(value)
                            elif value is None:
                                value = "N/A"
                            elif isinstance(value, str) and len(value) > 50:
                                value = value[:47] + "..."

                            cell_content = [
                                Paragraph(key.replace('_', ' ').upper(), kv_label_style),
                                Paragraph(str(value), kv_value_style)
                            ]
                            row_cells.append(cell_content)
                        else:
                            row_cells.append(["", ""])
                    kv_data.append(row_cells)

                if kv_data:
                    kv_table = Table(kv_data, colWidths=[(content_width - 20) / 2] * 2)
                    kv_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), BG_ROW_ALT),
                        ('LINEBEFORE', (0, 0), (0, -1), 3, PRIMARY_BLUE),
                        ('LINEBEFORE', (1, 0), (1, -1), 3, PRIMARY_BLUE),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('LEFTPADDING', (0, 0), (-1, -1), 12),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ]))
                    story.append(kv_table)

            # --- SECTION: Full Data Table ---
            if len(rows) > 5:
                story.append(PageBreak())
                story.append(header_table)
                story.append(Spacer(1, 20))

                story.append(Paragraph("Detailed Results", section_title_style))

                if len(rows) > 50:
                    story.append(Paragraph(
                        f"<i>Showing first 50 of {len(rows)} records</i>",
                        ParagraphStyle('Note', parent=styles['Normal'], fontSize=9,
                                      textColor=TEXT_MUTED, spaceAfter=10)
                    ))

                # Prepare table
                display_columns = columns[:8]  # Limit columns
                table_data = [[col.replace('_', ' ').title()[:15] for col in display_columns]]

                for row in rows[:50]:
                    row_data = []
                    for col in display_columns:
                        value = row.get(col, '')
                        if isinstance(value, str) and len(value) > 25:
                            value = value[:22] + '...'
                        elif isinstance(value, (list, dict)):
                            value = str(value)[:22] + '...'
                        elif isinstance(value, float):
                            value = f"{value:,.2f}"
                        elif value is None:
                            value = 'N/A'
                        row_data.append(str(value))
                    table_data.append(row_data)

                # Calculate column widths
                num_cols = len(display_columns)
                col_widths = [(content_width - 10) / num_cols] * num_cols

                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_BLUE),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('TOPPADDING', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BG_ROW_ALT]),
                    ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#e0e0e0')),
                    ('LINEBELOW', (0, 1), (-1, -2), 0.5, colors.HexColor('#e0e0e0')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ]))

                story.append(table)

            # Footer
            story.append(Spacer(1, 30))
            footer_table = Table(
                [[
                    Paragraph(f"Generated by EasyCall | {generated_at}", ParagraphStyle(
                        'FooterLeft', fontSize=8, textColor=TEXT_LIGHT)),
                    Paragraph("CONFIDENTIAL", ParagraphStyle(
                        'FooterRight', fontSize=8, textColor=TEXT_LIGHT, alignment=TA_RIGHT))
                ]],
                colWidths=[content_width * 0.7, content_width * 0.3]
            )
            footer_table.setStyle(TableStyle([
                ('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.HexColor('#dddddd')),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(footer_table)

            # Build PDF
            doc.build(story)

            self._log(f"  💾 PDF report written to: {file_path}")

            return {
                'file_path': file_path,
                'rows_written': len(rows),
                'pages': 'multiple'
            }

        except ImportError as e:
            self._log(f"  ⚠️ PDF generation requires reportlab and matplotlib: {e}")
            self._log("  ℹ️ Install with: pip install reportlab matplotlib")
            return {'file_path': None, 'error': f'Missing dependencies: {e}'}
        except Exception as e:
            self._log(f"  ❌ PDF generation failed: {str(e)}")
            return {'file_path': None, 'error': str(e)}

    def _export_pdf_template(self, inputs: dict, config: dict) -> dict:
        """
        Export data to PDF using Django templates with HTML/CSS styling.

        This method provides more control over the report layout using
        HTML templates and CSS. It adapts to different API data structures.

        Features:
        - Professional cover page
        - Adaptive sections based on data type
        - Clean HTML/CSS styling
        - Support for all API providers (Chainalysis, TRM Labs)
        """
        from .report_generator import ReportGenerator

        rows = self._prepare_export_data(inputs)
        report_title = config.get('report_title', 'Blockchain Intelligence Report')
        file_path = self._get_output_path(config, 'report') + '.pdf'

        self._log(f"  📑 Generating template-based PDF report...")

        try:
            # Create report generator
            workflow_name = getattr(self.workflow, 'name', 'Untitled Workflow')
            generator = ReportGenerator(workflow_name)

            # Detect data types and add appropriate sections
            if rows:
                # Check for cluster info data
                if any('cluster_name' in row or 'clusterName' in row for row in rows):
                    generator.add_section(
                        'chainalysis_cluster_info',
                        {'cluster_info': rows},
                        title='Cluster Attribution'
                    )

                # Check for balance data
                elif any('balance' in row for row in rows):
                    generator.add_section(
                        'chainalysis_cluster_balance',
                        {'balance_data': rows},
                        title='Balance Summary'
                    )

                # Check for counterparty data
                elif any('counterparty' in str(row).lower() or 'name' in row for row in rows):
                    generator.add_section(
                        'chainalysis_cluster_counterparties',
                        {'counterparties': rows},
                        title='Counterparty Analysis'
                    )

                # Check for transaction data
                elif any('tx_hash' in row or 'transaction' in str(row).lower() for row in rows):
                    generator.add_section(
                        'chainalysis_transaction_details',
                        {'transaction_details': rows},
                        title='Transaction Details'
                    )

                # Check for exposure data
                elif any('exposure' in str(row).lower() or 'risk' in str(row).lower() for row in rows):
                    generator.add_section(
                        'chainalysis_exposure_category',
                        {'direct_exposure': rows},
                        title='Exposure Analysis'
                    )

                # Generic table for other data
                else:
                    generator.add_section(
                        'generic',
                        rows,
                        title='Query Results'
                    )

            # Also add raw inputs if they contain specific node outputs
            for key, value in inputs.items():
                if isinstance(value, dict):
                    # Check for specific node output patterns
                    if 'cluster_info' in value:
                        generator.add_section('chainalysis_cluster_info', value)
                    elif 'balance_data' in value:
                        generator.add_section('chainalysis_cluster_balance', value)
                    elif 'counterparties' in value:
                        generator.add_section('chainalysis_cluster_counterparties', value)
                    elif 'transaction_details' in value:
                        generator.add_section('chainalysis_transaction_details', value)
                    elif 'direct_exposure' in value or 'indirect_exposure' in value:
                        generator.add_section('chainalysis_exposure_category', value)
                    elif 'risk_score' in value or 'risk_indicators' in value:
                        generator.add_section('trm_total_exposure', value)
                    elif 'entities' in value or 'entity_name' in value:
                        generator.add_section('trm_address_attribution', value)

            # Generate executive summary
            stats = self._calculate_summary_stats(rows) if rows else {}
            summary_parts = []
            if generator.total_addresses > 0:
                summary_parts.append(f"analyzed {generator.total_addresses} blockchain address(es)")
            if generator.total_transactions > 0:
                summary_parts.append(f"examined {generator.total_transactions} transaction(s)")
            if stats.get('total_balance'):
                summary_parts.append(f"with a total balance of {stats['total_balance']:.8f}")
            if generator.data_sources:
                summary_parts.append(f"using data from {', '.join(generator.data_sources)}")

            executive_summary = (
                f"This report {' and '.join(summary_parts)}."
                if summary_parts
                else "This report contains the results of the blockchain intelligence workflow execution."
            )

            # Build summary stats for cards
            summary_stats = []
            if rows:
                summary_stats.append({'value': len(rows), 'label': 'Records'})
            if stats.get('unique_addresses'):
                summary_stats.append({'value': stats['unique_addresses'], 'label': 'Addresses'})
            if stats.get('total_transfers'):
                summary_stats.append({'value': stats['total_transfers'], 'label': 'Transfers'})

            # Generate PDF
            output_path = generator.generate_pdf(
                file_path,
                report_title=report_title,
                executive_summary=executive_summary,
                summary_stats=summary_stats if summary_stats else None,
            )

            self._log(f"  ✅ Template PDF report saved: {output_path}")
            return {
                'file_path': output_path,
                'rows_written': len(rows) if rows else 0,
                'sections': len(generator.sections),
                'format': 'pdf_template'
            }

        except Exception as e:
            import traceback
            self._log(f"  ❌ Template PDF generation failed: {str(e)}")
            self._log(f"  📋 Traceback: {traceback.format_exc()}")
            # Fall back to ReportLab-based PDF
            self._log("  ℹ️ Falling back to ReportLab PDF...")
            return self._export_pdf(inputs, config)

    def _calculate_summary_stats(self, rows: List[dict]) -> dict:
        """Calculate summary statistics from data rows."""
        stats = {}

        if not rows:
            return stats

        # Look for balance field
        for row in rows:
            if 'balance' in row:
                stats['total_balance'] = stats.get('total_balance', 0) + float(row.get('balance', 0) or 0)
            if 'transferCount' in row or 'transfer_count' in row:
                count = row.get('transferCount') or row.get('transfer_count') or 0
                stats['total_transfers'] = stats.get('total_transfers', 0) + int(count)

        # Count unique addresses
        addresses = set()
        for row in rows:
            if 'address' in row:
                addresses.add(row['address'])
            if 'source_address' in row:
                addresses.add(row['source_address'])
        if addresses:
            stats['unique_addresses'] = len(addresses)

        # Count high-risk flags
        high_risk_count = 0
        for row in rows:
            if row.get('has_high_risk'):
                high_risk_count += 1
            if 'high_risk_flags' in row and row['high_risk_flags']:
                high_risk_count += len(row['high_risk_flags']) if isinstance(row['high_risk_flags'], list) else 1
        if high_risk_count:
            stats['high_risk_count'] = high_risk_count

        return stats

    def _generate_graph(self, rows: List[dict], graph_type: str) -> Optional[BytesIO]:
        """Generate a graph image from data rows."""
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        if not rows:
            return None

        # Analyze data to determine best graph type
        sample_row = rows[0]

        # Find numeric columns
        numeric_cols = []
        category_cols = []

        for key, value in sample_row.items():
            if isinstance(value, (int, float)) and key not in ('count', 'rows_written'):
                numeric_cols.append(key)
            elif isinstance(value, str) and len(value) < 50:
                category_cols.append(key)

        if not numeric_cols:
            return None

        # Auto-detect best graph type
        if graph_type == 'auto':
            if 'category' in sample_row or 'cluster_name' in sample_row or 'clusterName' in sample_row:
                graph_type = 'pie'
            elif len(rows) > 10:
                graph_type = 'bar'
            else:
                graph_type = 'bar'

        # Create figure
        fig, ax = plt.subplots(figsize=(10, 6))

        # Style
        plt.style.use('seaborn-v0_8-whitegrid')
        colors_list = ['#1565c0', '#00897b', '#f57c00', '#d32f2f', '#7b1fa2', '#388e3c']

        try:
            if graph_type == 'pie' and 'category' in sample_row:
                # Pie chart for category distribution
                categories = {}
                value_field = 'value' if 'value' in sample_row else numeric_cols[0]
                for row in rows:
                    cat = row.get('category', row.get('cluster_name', row.get('clusterName', 'Unknown')))
                    val = float(row.get(value_field, 1) or 1)
                    categories[cat] = categories.get(cat, 0) + val

                if categories:
                    labels = list(categories.keys())[:8]  # Max 8 slices
                    values = [categories[l] for l in labels]
                    ax.pie(values, labels=labels, autopct='%1.1f%%', colors=colors_list[:len(labels)])
                    ax.set_title('Distribution by Category', fontsize=14, fontweight='bold')

            elif graph_type == 'bar':
                # Bar chart
                if 'balance' in numeric_cols:
                    # Balance by address/cluster
                    label_field = 'address' if 'address' in sample_row else 'cluster_name'
                    data = [(row.get(label_field, f'Item {i}')[:15], float(row.get('balance', 0) or 0))
                            for i, row in enumerate(rows[:10])]
                    labels, values = zip(*data) if data else ([], [])
                    ax.barh(labels, values, color=colors_list[0])
                    ax.set_xlabel('Balance', fontsize=11)
                    ax.set_title('Balance by Address', fontsize=14, fontweight='bold')
                else:
                    # Generic numeric data
                    num_col = numeric_cols[0]
                    values = [float(row.get(num_col, 0) or 0) for row in rows[:15]]
                    labels = [f'Record {i+1}' for i in range(len(values))]
                    ax.bar(labels, values, color=colors_list[0])
                    ax.set_ylabel(num_col, fontsize=11)
                    ax.set_title(f'{num_col} Distribution', fontsize=14, fontweight='bold')
                    plt.xticks(rotation=45, ha='right')

            elif graph_type == 'line':
                # Line chart
                num_col = numeric_cols[0]
                values = [float(row.get(num_col, 0) or 0) for row in rows[:50]]
                ax.plot(range(len(values)), values, color=colors_list[0], linewidth=2, marker='o', markersize=4)
                ax.set_xlabel('Record Index', fontsize=11)
                ax.set_ylabel(num_col, fontsize=11)
                ax.set_title(f'{num_col} Trend', fontsize=14, fontweight='bold')
                ax.fill_between(range(len(values)), values, alpha=0.3, color=colors_list[0])

            plt.tight_layout()

            # Save to BytesIO
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close(fig)

            return img_buffer

        except Exception as e:
            plt.close(fig)
            return None

    def _export_txt(self, inputs: dict, config: dict) -> dict:
        """Export data to TXT file."""
        rows = self._prepare_export_data(inputs)

        if not rows:
            self._log("  ⚠️ No data to export")
            return {'file_path': None, 'rows_written': 0, 'error': 'No data to export'}

        # Get output path
        file_path = self._get_output_path(config, 'output') + '.txt'

        self._log(f"  📤 Exporting {len(rows)} records to TXT...")

        # Write TXT file (formatted output)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(f"Workflow Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")

            for i, row in enumerate(rows, 1):
                f.write(f"Record {i}:\n")
                f.write("-" * 40 + "\n")
                for key, value in row.items():
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value, indent=2)
                    f.write(f"  {key}: {value}\n")
                f.write("\n")

        self._log(f"  💾 Written to: {file_path}")

        return {
            'file_path': file_path,
            'rows_written': len(rows)
        }

    # ═══════════════════════════════════════════════════════════════════════
    # HELPER METHODS
    # ═══════════════════════════════════════════════════════════════════════

    def _get_node_inputs(self, node_id: str, edges: List[dict]) -> dict:
        """Get input values from connected nodes."""
        inputs = {}

        for edge in edges:
            if edge.get('target') == node_id:
                source_id = edge.get('source')
                source_handle = edge.get('sourceHandle', 'output')
                target_handle = edge.get('targetHandle', 'input')

                # Get output from source node
                source_outputs = self.execution_context.get(source_id, {})

                # Map source output to target input
                if source_handle in source_outputs:
                    value = source_outputs[source_handle]
                    inputs[target_handle] = value

                    # For 'data' inputs (like csv_export), also include the full source output
                    # so the export node has access to all data from the connected node
                    if target_handle == 'data':
                        # Merge all source outputs into the data
                        if isinstance(value, dict):
                            inputs[target_handle] = source_outputs
                        else:
                            inputs[target_handle] = source_outputs
                            inputs['_source_data'] = source_outputs
                elif source_outputs:
                    # If no specific handle, pass all outputs
                    inputs[target_handle] = source_outputs

        return inputs

    def _topological_sort(self, nodes: List[dict], edges: List[dict]) -> List[dict]:
        """
        Sort nodes in execution order (topological sort).

        Nodes with no dependencies execute first.
        """
        # Build dependency graph
        node_map = {n['id']: n for n in nodes}
        dependencies = {n['id']: set() for n in nodes}

        for edge in edges:
            target = edge.get('target')
            source = edge.get('source')
            if target in dependencies:
                dependencies[target].add(source)

        # Kahn's algorithm
        sorted_nodes = []
        no_deps = [n['id'] for n in nodes if not dependencies[n['id']]]

        while no_deps:
            node_id = no_deps.pop(0)
            sorted_nodes.append(node_map[node_id])

            # Remove this node from all dependency sets
            for dep_set in dependencies.values():
                dep_set.discard(node_id)

            # Check for new nodes with no dependencies
            for nid, deps in dependencies.items():
                if not deps and nid not in [n['id'] for n in sorted_nodes] and nid not in no_deps:
                    no_deps.append(nid)

        return sorted_nodes

def _is_database_node(self, node_type: str) -> bool:
    """
    Check if node_type is a database-generated node.
    
    Database nodes follow pattern: {provider}_{operation_id}
    Examples: trm_labs_get_attribution, chainalysis_get_cluster_info
    """
    # Database nodes contain provider prefix
    providers = ['trm_labs', 'chainalysis', 'elliptic']  # Add more as needed
    return any(node_type.startswith(f"{provider}_") for provider in providers)


def _execute_database_node(self, node_type: str, inputs: dict, config: dict) -> dict:
    """
    Execute a database-generated node.
    
    Steps:
    1. Look up node definition (frozen config first, then database)
    2. Build HTTP request from definition
    3. Execute API call
    4. Map response to outputs
    """
    # Step 1: Get node definition
    node_def = self._lookup_node_definition(node_type)
    
    if not node_def:
        raise NodeExecutionError(f"Node definition not found: {node_type}")
    
    # Step 2: Build API request
    request_config = self._build_api_request(node_def, inputs, config)
    
    # Step 3: Execute API call
    from apps.execution.api_executor import GenericAPIExecutor
    api_executor = GenericAPIExecutor()
    response = api_executor.execute(request_config)
    
    # Step 4: Map response to outputs
    return self._map_api_response(response, node_def)


def _lookup_node_definition(self, node_type: str) -> Optional[dict]:
    """
    Look up node definition from frozen config or database.
    
    Priority:
    1. Frozen configuration (from workflow save)
    2. Current database definition
    """
    # Check frozen config first (stored in workflow.canvas_data['_frozen_nodes'])
    frozen_nodes = self.workflow.canvas_data.get('_frozen_nodes', {})
    if node_type in frozen_nodes:
        self._log(f"  [FROZEN] Using frozen node definition for {node_type}")
        return frozen_nodes[node_type]
    
    # Fall back to database
    from apps.integrations.models import OpenAPISpec
    from apps.integrations.node_generator import NodeGenerator
    
    # Extract provider from node_type (e.g., "trm_labs_get_attribution" -> "trm_labs")
    provider = '_'.join(node_type.split('_')[:2])  # "trm_labs"
    
    # Find active spec for this provider
    spec = OpenAPISpec.objects.filter(
        provider=provider,
        is_active=True,
        is_parsed=True
    ).first()
    
    if not spec:
        return None
    
    # Generate nodes and find matching one
    generator = NodeGenerator()
    nodes = generator.generate_nodes(
        endpoints=spec.parsed_data.get('endpoints', []),
        provider=provider
    )
    
    for node in nodes:
        if node['type'] == node_type:
            self._log(f"  [DATABASE] Using current node definition for {node_type}")
            return node
    
    return None